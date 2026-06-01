"""Parser robuste des relevés HSBC (XLSX + PDF) -> list[BankTransaction].
NEUTRE : rend toutes les lignes de mouvement, ne filtre pas le métier (interne/frais).
Le filtrage est fait dans service/bank_match.

Colonnes XLSX attendues : Libellé, Référence client, Montant du crédit, Montant du débit,
Date de valeur, Date opération, Référence bancaire. En-têtes de compte répétés -> ignorés.
Normaliser l'encodage (remplacer les '?' parasites type 'Tesla?FR?'). Dates DD/MM/YYYY.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from core.models import BankTransaction

logger = logging.getLogger(__name__)

# --- Détection des colonnes ------------------------------------------------
# On associe chaque champ de BankTransaction à un (ou plusieurs) en-tête(s)
# HSBC possible(s). La comparaison se fait sur une forme normalisée
# (minuscule, sans accent, espaces compactés) pour tolérer les variantes.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "label": ("libelle",),
    "client_ref": ("reference client",),
    "credit": ("montant du credit", "credit"),
    "debit": ("montant du debit", "debit"),
    "value_date": ("date de valeur",),
    "op_date": ("date operation",),
    "ref": ("reference bancaire",),
}


def _strip_accents(value: str) -> str:
    """Retire les diacritiques pour comparer les en-têtes de façon tolérante."""
    decomposed = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def _normalize_header(value: object) -> str:
    """Normalise un en-tête : minuscule, sans accent, espaces compactés."""
    text = _strip_accents(str(value or "")).lower()
    return re.sub(r"\s+", " ", text).strip()


def _normalize_text(value: object) -> str:
    """Normalise un libellé : enlève les '?' parasites (ex 'Tesla?FR?'),
    compacte les espaces. Garde les accents (lisibilité métier)."""
    if value is None:
        return ""
    text = str(value).replace("?", " ")
    return re.sub(r"\s+", " ", text).strip()


def _parse_date(value: object) -> date | None:
    """Parse une date HSBC. Accepte un datetime/date natif (cellule Excel typée)
    ou une chaîne au format DD/MM/YYYY (avec - ou . comme séparateurs)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: object) -> Decimal | None:
    """Parse un montant HSBC -> Decimal. Accepte un nombre natif ou une chaîne
    au format français ('1 234,56', '1.234,56', espaces insécables). Retourne
    None si la cellule est vide ou non numérique."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    if not text:
        return None
    # Retire espaces (y compris insécables) et symboles monétaires.
    text = text.replace("\xa0", "").replace(" ", "").replace("€", "").replace("EUR", "")
    if not text:
        return None
    # Format français : '.' = séparateur de milliers, ',' = décimale.
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _build_transaction(row: dict[str, object]) -> BankTransaction | None:
    """Construit une BankTransaction depuis un dict {champ: valeur brute}.
    Retourne None si la ligne n'est pas un mouvement (pas de date exploitable,
    ou ni crédit ni débit) : c'est ainsi qu'on ignore les en-têtes de compte
    répétés (IBAN/BIC/soldes), sans filtrage métier."""
    value_date = _parse_date(row.get("value_date"))
    op_date = _parse_date(row.get("op_date"))
    credit = _parse_amount(row.get("credit"))
    debit = _parse_amount(row.get("debit"))

    # Une ligne de mouvement a au moins une date et au moins un montant.
    if value_date is None and op_date is None:
        return None
    if credit is None and debit is None:
        return None

    # value_date est obligatoire dans le modèle : on retombe sur op_date.
    if value_date is None:
        value_date = op_date  # type: ignore[assignment]

    label = _normalize_text(row.get("label"))
    client_ref = _normalize_text(row.get("client_ref")) or None
    ref = _normalize_text(row.get("ref"))

    return BankTransaction(
        ref=ref,
        value_date=value_date,
        op_date=op_date,
        label=label,
        client_ref=client_ref,
        credit=credit,
        debit=debit,
        source="hsbc",
    )


def _map_header_row(cells: list[object]) -> dict[str, int] | None:
    """Tente d'interpréter une ligne comme en-tête : renvoie {champ: index} si
    au moins une colonne date ET une colonne montant sont présentes, sinon None."""
    normalized = [_normalize_header(c) for c in cells]
    mapping: dict[str, int] = {}
    for field, aliases in _COLUMN_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases and field not in mapping:
                mapping[field] = idx
                break
    has_date = "value_date" in mapping or "op_date" in mapping
    has_amount = "credit" in mapping or "debit" in mapping
    if has_date and has_amount:
        return mapping
    return None


def _rows_to_transactions(rows: list[list[object]]) -> list[BankTransaction]:
    """Logique commune XLSX/PDF : trouve la ligne d'en-tête, mappe les colonnes,
    puis convertit chaque ligne en BankTransaction (best-effort, ne plante pas)."""
    header_map: dict[str, int] | None = None
    transactions: list[BankTransaction] = []

    for cells in rows:
        if not cells:
            continue
        if header_map is None:
            header_map = _map_header_row(cells)
            continue
        row = {
            field: (cells[idx] if idx < len(cells) else None)
            for field, idx in header_map.items()
        }
        try:
            tx = _build_transaction(row)
        except Exception:  # best-effort : une ligne ambiguë ne casse pas le tout
            logger.warning("Ligne HSBC ignorée (parsing impossible): %r", cells)
            continue
        if tx is not None:
            transactions.append(tx)

    return transactions


def parse_hsbc_xlsx(file_bytes: bytes) -> list[BankTransaction]:
    """Parse un export Excel HSBC -> list[BankTransaction].

    Colonnes attendues : Libellé, Référence client, Montant du crédit,
    Montant du débit, Date de valeur, Date opération, Référence bancaire.
    Les en-têtes de compte répétés (IBAN/BIC/soldes) sont ignorés. NEUTRE :
    aucun filtrage métier (internes/frais inclus dans le retour).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    rows: list[list[object]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    wb.close()
    return _rows_to_transactions(rows)


def parse_hsbc_pdf(file_bytes: bytes) -> list[BankTransaction]:
    """Parse un relevé PDF HSBC via pdfplumber (extraction de tableaux).
    Best-effort robuste : si une ligne est ambiguë, on la saute sans planter."""
    import pdfplumber

    rows: list[list[object]] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            try:
                tables = page.extract_tables()
            except Exception:
                logger.warning("Extraction de table échouée sur une page PDF HSBC")
                continue
            for table in tables or []:
                for row in table:
                    rows.append(list(row))
    return _rows_to_transactions(rows)


def parse_hsbc(file_bytes: bytes, filename: str) -> list[BankTransaction]:
    """Dispatch selon extension (.xlsx/.xls -> xlsx ; .pdf -> pdf)."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        return parse_hsbc_xlsx(file_bytes)
    if name.endswith(".pdf"):
        return parse_hsbc_pdf(file_bytes)
    raise ValueError(f"Format de fichier HSBC non supporté : {filename!r}")
